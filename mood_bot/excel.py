import io
import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


def _fmt_time_cell(created_at: str, end_time: str | None) -> str:
    """Возвращает '11:00–16:00' если есть end_time, иначе 'HH:MM'."""
    start = created_at.split(" ")[1][:5]
    if end_time:
        return f"{start}–{end_time[:5]}"
    return start


def build_abc_excel(rows: list[sqlite3.Row]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "КПТ ABC"

    headers = ["Дата", "Время", "A — Ситуация", "B — Мысли/убеждения", "C — Чувства/эмоции", "Комментарий"]
    header_fill = PatternFill("solid", fgColor="7B68EE")
    header_font = Font(bold=True, color="FFFFFF")

    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for row_idx, row in enumerate(rows, start=2):
        iso_date = row["created_at"].split(" ")[0]
        date_part = f"{iso_date[8:10]}.{iso_date[5:7]}.{iso_date[0:4]}"
        time_part = row["created_at"].split(" ")[1][:5]

        ws.cell(row=row_idx, column=1, value=date_part)
        ws.cell(row=row_idx, column=2, value=time_part).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=3, value=row["situation"]).alignment = Alignment(wrap_text=True)
        ws.cell(row=row_idx, column=4, value=row["thoughts"]).alignment = Alignment(wrap_text=True)
        ws.cell(row=row_idx, column=5, value=row["feelings"]).alignment = Alignment(wrap_text=True)
        ws.cell(row=row_idx, column=6, value=row["comment"] or "").alignment = Alignment(wrap_text=True)

    widths = [12, 8, 35, 35, 35, 30]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_excel(rows: list[sqlite3.Row], total: int) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Настроение"

    # ── Заголовки ──────────────────────────────────────────────
    headers = ["Дата", "Время", "Отправлено", "Описание", "Оценка"]
    header_fill = PatternFill("solid", fgColor="4F81BD")
    header_font = Font(bold=True, color="FFFFFF")

    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # ── Данные ─────────────────────────────────────────────────
    for row_idx, row in enumerate(rows, start=2):
        created_at: str = row["created_at"]
        iso_date = created_at.split(" ")[0]
        date_part = f"{iso_date[8:10]}.{iso_date[5:7]}.{iso_date[0:4]}"
        time_cell_value = _fmt_time_cell(created_at, row["end_time"])

        # Время отправки — берём из sent_at, если есть, иначе из created_at
        sent_at: str | None = row["sent_at"]
        sent_time = sent_at.split(" ")[1][:5] if sent_at else created_at.split(" ")[1][:5]

        ws.cell(row=row_idx, column=1, value=date_part)
        ws.cell(row=row_idx, column=2, value=time_cell_value).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=3, value=sent_time).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=4, value=row["description"])
        score_cell = ws.cell(row=row_idx, column=5, value=row["score"])
        score_cell.alignment = Alignment(horizontal="center")

        if row["score"] > 0:
            score_cell.font = Font(color="375623", bold=True)
        elif row["score"] < 0:
            score_cell.font = Font(color="9C0006", bold=True)
        else:
            score_cell.font = Font(color="0070C0", bold=True)

    # ── Итоговая строка ────────────────────────────────────────
    total_row = len(rows) + 2
    ws.cell(row=total_row, column=4, value="ИТОГО").font = Font(bold=True)
    total_cell = ws.cell(row=total_row, column=5, value=total)
    total_cell.font = Font(bold=True)
    total_cell.alignment = Alignment(horizontal="center")

    # ── Ширина колонок ─────────────────────────────────────────
    widths = [12, 14, 12, 40, 8]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
