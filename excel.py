import io
import re
import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

_RANGE_RE = re.compile(
    r"с\s+(\d{1,2})(?::(\d{2}))?\s+до\s+(\d{1,2})(?::(\d{2}))?",
    re.IGNORECASE,
)


def _fmt_time_cell(created_at: str, description: str) -> str:
    """Возвращает '11:00–16:00' если в описании есть диапазон, иначе 'HH:MM'."""
    m = _RANGE_RE.search(description)
    if m:
        start_h, start_m = int(m.group(1)), int(m.group(2) or 0)
        end_h, end_m = int(m.group(3)), int(m.group(4) or 0)
        return f"{start_h:02d}:{start_m:02d}–{end_h:02d}:{end_m:02d}"
    return created_at.split(" ")[1][:5]  # "HH:MM"


def build_excel(rows: list[sqlite3.Row], total: int) -> io.BytesIO:
    """
    Принимает строки из БД (created_at, score, description) и итоговую сумму.
    Возвращает BytesIO с .xlsx-файлом.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Настроение"

    # ── Заголовки ──────────────────────────────────────────────
    headers = ["Дата", "Время", "Описание", "Оценка"]
    header_fill = PatternFill("solid", fgColor="4F81BD")
    header_font = Font(bold=True, color="FFFFFF")

    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # ── Данные ─────────────────────────────────────────────────
    for row_idx, row in enumerate(rows, start=2):
        created_at: str = row["created_at"]
        date_part = created_at.split(" ")[0]
        time_cell_value = _fmt_time_cell(created_at, row["description"])

        ws.cell(row=row_idx, column=1, value=date_part)
        ws.cell(row=row_idx, column=2, value=time_cell_value)
        ws.cell(row=row_idx, column=3, value=row["description"])
        score_cell = ws.cell(row=row_idx, column=4, value=row["score"])
        score_cell.alignment = Alignment(horizontal="center")

        # Цветовая индикация оценки
        if row["score"] > 0:
            score_cell.font = Font(color="375623")
        elif row["score"] < 0:
            score_cell.font = Font(color="9C0006")

    # ── Итоговая строка ────────────────────────────────────────
    total_row = len(rows) + 2
    ws.cell(row=total_row, column=3, value="ИТОГО").font = Font(bold=True)
    total_cell = ws.cell(row=total_row, column=4, value=total)
    total_cell.font = Font(bold=True)
    total_cell.alignment = Alignment(horizontal="center")

    # ── Ширина колонок ─────────────────────────────────────────
    widths = [12, 14, 40, 8]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
